import streamlit as st
import pandas as pd
import numpy as np
from PIL import Image
from datetime import datetime, timedelta
import os, pickle, re
from streamlit import rerun
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
import pydeck as pdk
import plotly.express as px
import plotly.graph_objects as go

with open("style.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

# ---------------------- CONFIG ----------------------
st.set_page_config(page_title="Base de datos Radiaciones No Ionizantes ENACOM", layout="wide")

# Logo ENACOM
try:
    logo = Image.open("logo_enacom.png")
    st.sidebar.image(logo, width=200)
except:
    st.sidebar.write("Logo ENACOM no encontrado")

# ------------------- SESSION STATE ------------------
st.session_state.setdefault("tabla_maestra", pd.DataFrame())
st.session_state.setdefault("uploaded_files_list", [])
st.session_state.setdefault("form_ccte", "")
st.session_state.setdefault("form_provincia", "")
st.session_state.setdefault("form_localidad", "")
st.session_state.setdefault("form_expediente", "")

# Cargar tabla maestra persistida
PICKLE_FILE = "tabla_maestra.pkl"
if st.session_state["tabla_maestra"].empty and os.path.exists(PICKLE_FILE):
    try:
        st.session_state["tabla_maestra"] = pickle.load(open(PICKLE_FILE, "rb"))
    except Exception as e:
        st.warning(f"No se pudo cargar {PICKLE_FILE}: {e}")

# ------------------- FUNCIONES ----------------------
def format_timedelta_long(td: timedelta) -> str:
    total_seconds = int(td.total_seconds())
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def parse_dms_to_decimal(val):
    if pd.isna(val):
        return np.nan
    try:
        return float(val)
    except:
        pass
    s = str(val).strip().replace(",", ".")
    m = re.search(r'([+-]?\d+(?:\.\d+)?)\D+(\d+(?:\.\d+)?)\D+(\d+(?:\.\d+)?)\D*\s*([NnSsEeWwOo])?', s)
    if m:
        d = float(m.group(1)); mnt = float(m.group(2)); sec = float(m.group(3))
        hemi = (m.group(4) or "").upper()
        dec = abs(d) + mnt/60.0 + sec/3600.0
        if hemi in ("S","W","O"):
            dec = -dec
        return dec
    m2 = re.search(r'([-+]?\d+(?:\.\d+)?)', s)
    if m2:
        try:
            return float(m2.group(1))
        except:
            return np.nan
    return np.nan

def extract_numeric_from_text(series):
    s = series.astype(str).str.replace(",", ".", regex=False)
    num = s.str.extract(r'([-+]?\d+(?:\.\d+)?(?:[eE][-+]?\d+)?)', expand=False)
    return pd.to_numeric(num, errors="coerce")

def find_index_column(df):
    candidates = ["índice","indice","index","nro","nº","n°","num","numero","#"]
    for c in df.columns:
        low = str(c).lower()
        for cand in candidates:
            if cand in low:
                return c
    return None

#Fecha Hora - cálculo
def calcular_tiempo_total_por_archivo(df: pd.DataFrame) -> timedelta:
    total = timedelta(0)
    if "Nombre Archivo" in df.columns:
        for archivo, df_archivo in df.groupby("Nombre Archivo"):
            if "Fecha" in df_archivo.columns and "Hora" in df_archivo.columns:
                # Cada día dentro del archivo
                for fecha, df_dia in df_archivo.groupby("Fecha"):
                    horas_validas = df_dia["Hora"].dropna()
                    if len(horas_validas) > 0:
                        dt_inicio = datetime.combine(fecha, horas_validas.min())
                        dt_fin = datetime.combine(fecha, horas_validas.max())
                        delta = dt_fin - dt_inicio
                        if delta.total_seconds() < 0:
                            delta += timedelta(days=1)
                        total += delta
    return total

def procesar_archivos(uploaded_files, ccte, provincia, localidad, expediente):
    lista_procesados = []
    resumen_archivos = []
    for file in uploaded_files:
        try:
            df = pd.read_excel(file, header=8, engine="openpyxl")
        except Exception as e:
            st.warning(f"No se pudo leer {file.name}: {e}")
            continue

        df = df.dropna(axis=1, how="all")
        idx_col = find_index_column(df)
        total_mediciones = len(df)
        if idx_col:
            df["_idx_num"] = pd.to_numeric(df[idx_col], errors="coerce")
            df = df[df["_idx_num"].notna()]
            if not df.empty:
                total_mediciones = int(df["_idx_num"].max())
        mapping_candidates = {
            "Fecha": ["fecha"],
            "Hora": ["hora", "time"],
            "Resultado": ["resultado con incertidumbre","resultado"],
            "Sonda": ["sonda","sonda utilizada"],
            "Lat": ["latitud","lat"],
            "Lon": ["longitud","lon"]
        }
        columnas_map = {}
        missing = False
        for key, cands in mapping_candidates.items():
            found = None
            for c in df.columns:
                low = str(c).lower()
                for cand in cands:
                    if cand in low:
                        found = c
                        break
                if found:
                    break
            if not found and key not in ("Lat","Lon"):
                st.warning(f"Archivo {file.name}: no se encontró columna para '{key}'")
                missing = True
                break
            if found:
                columnas_map[key] = found
        if missing:
            continue

        df = df.rename(columns={v:k for k,v in columnas_map.items()})
        df["CCTE"] = ccte
        df["Provincia"] = provincia
        df["Localidad"] = localidad
        df["Expediente"] = expediente if expediente else os.path.splitext(file.name)[0]
        df["Nombre Archivo"] = file.name
        
        #Resultado
        if "Resultado" in df.columns:
            df["Resultado"] = extract_numeric_from_text(df["Resultado"])
        if "Lat" in df.columns:
            df["Lat"] = df["Lat"].apply(parse_dms_to_decimal)
        if "Lon" in df.columns:
            df["Lon"] = df["Lon"].apply(parse_dms_to_decimal)
        if "_idx_num" in df.columns:
            df = df.drop(columns=["_idx_num"])

        lista_procesados.append(df)
        max_res = df["Resultado"].max() if "Resultado" in df.columns else None
        resumen_archivos.append({
            "archivo": file.name,
            "expediente": df["Expediente"].iloc[0],
            "total mediciones": total_mediciones,
            "max_resultado": max_res
        })

    if lista_procesados:
        return pd.concat(lista_procesados, ignore_index=True), pd.DataFrame(resumen_archivos)
    return pd.DataFrame(), pd.DataFrame()

# ---------------- Función: eliminar localidad ----------------
def eliminar_localidad(nombre_localidad: str):
   
    if "tabla_maestra" not in st.session_state or st.session_state["tabla_maestra"].empty:
        st.warning("⚠️ No hay datos cargados en la tabla maestra.")
        return

    df = st.session_state["tabla_maestra"]
    if "Localidad" not in df.columns:
        st.error("❌ La tabla no tiene columna 'Localidad'.")
        return

    # Cantidad antes
    cantidad_antes = len(df)

    # Filtrar quitando la localidad
    df_filtrado = df[df["Localidad"] != nombre_localidad].copy()

    # Cantidad eliminada
    cantidad_despues = len(df_filtrado)
    eliminados = cantidad_antes - cantidad_despues

    if eliminados == 0:
        st.info(f"ℹ️ No se encontró la localidad **{nombre_localidad}** en la tabla.")
        return

    # Actualizar session_state y guardar
    st.session_state["tabla_maestra"] = df_filtrado

    try:
        pickle.dump(df_filtrado, open(PICKLE_FILE, "wb"))
        st.success(f"✅ Localidad **{nombre_localidad}** eliminada ({eliminados} registros).")
    except Exception as e:
        st.warning(f"⚠️ Localidad eliminada pero no se pudo guardar el pickle: {e}")


# ------------------- SIDEBAR: carga ------------------
st.sidebar.header("Cargar archivos")
if "uploader_key" not in st.session_state:
    st.session_state["uploader_key"] = 0

form_container = st.sidebar.empty()

with form_container.form("carga_form", clear_on_submit=False):
    ccte = st.selectbox("CCTE", ["CABA", "Buenos Aires", "Comodoro Rivadavia", "Córdoba", "Neuquén", "Posadas", "Salta"],
                        key="form_ccte")
    provincia = st.selectbox(
        "Provincia",
        ["Buenos Aires","CABA","Catamarca","Chaco","Chubut","Córdoba","Corrientes","Entre Ríos","Formosa","Jujuy",
         "La Pampa","La Rioja","Mendoza","Misiones","Neuquén","Río Negro","Salta","San Juan","San Luis","Santa Cruz",
         "Santa Fe","Santiago del Estero","Tierra del Fuego","Tucumán"],
        key="form_provincia"
    )
    localidad = st.text_input("Localidad", value=st.session_state["form_localidad"], key="form_localidad")
    expediente = st.text_input("Expediente", value=st.session_state["form_expediente"], key="form_expediente")

   
    files = st.file_uploader(
        "Seleccionar archivos Excel",
        accept_multiple_files=True,
        type=["xlsx"],
        key=f"form_files_{st.session_state['uploader_key']}"
    )
    submit = st.form_submit_button("Procesar archivos")

    if submit and files:
        df_proc, resumen_df = procesar_archivos(files, ccte, provincia, localidad, expediente)
        if not df_proc.empty:
            df_proc["FechaCarga"] = datetime.now()
            st.session_state["tabla_maestra"] = pd.concat([st.session_state["tabla_maestra"], df_proc], ignore_index=True)
            try:
                pickle.dump(st.session_state["tabla_maestra"], open(PICKLE_FILE,"wb"))
            except Exception as e:
                st.warning(f"No se pudo guardar pickle: {e}")
            st.session_state["uploaded_files_list"] = [f.name for f in files]
            st.success(f"{len(files)} archivos procesados y agregados a la tabla maestra.")
            st.sidebar.subheader("Resumen archivos procesados")
            st.sidebar.dataframe(resumen_df)
            st.session_state["uploader_key"] += 1
        else:
            st.warning("No se procesaron archivos. Revisá los avisos arriba.")

# Botón de restablecer
def reset_form():
    st.session_state["uploaded_files_list"] = []
    st.session_state["form_localidad"] = ""
    st.session_state["form_expediente"] = ""
    st.session_state["form_ccte"] = ""
    st.session_state["form_provincia"] = ""
    st.session_state["uploader_key"] += 1 
    rerun()

st.sidebar.button("Restablecer", on_click=reset_form)

# ------------------- SIDEBAR: eliminar localidad ------------------
if "tabla_maestra" in st.session_state and not st.session_state["tabla_maestra"].empty:
    localidades_unicas = sorted(st.session_state["tabla_maestra"]["Localidad"].dropna().unique().tolist())
    localidad_a_borrar = st.sidebar.selectbox("Seleccionar localidad a eliminar", [""] + localidades_unicas)

    if st.sidebar.button("❌ Eliminar localidad") and localidad_a_borrar:
        eliminar_localidad(localidad_a_borrar)

# ------------------- ENCABEZADO CON LOGO ------------------
col1, col2 = st.columns([6,1])

with col1:
    st.title(" ")

with col2:
    st.image("logo_enacom.png")

# ------------------- HIGHLIGHT GLOBAL ------------------
if "tabla_maestra" in st.session_state and not st.session_state["tabla_maestra"].empty:
    df = st.session_state["tabla_maestra"].copy()
    df["Resultado"] = pd.to_numeric(df["Resultado"], errors="coerce")
    idx_max = df["Resultado"].idxmax()
    fila_max = df.loc[idx_max]

    localidad_top = fila_max.get("Localidad", "N/A")
    resultado_top = fila_max["Resultado"]
    resultado_top_pct = resultado_top**2 / 3770 / 0.20021 * 100 if pd.notna(resultado_top) else None

    # Fecha/Hora asociada
    fecha_top = None
    if "FechaHora" in fila_max and pd.notna(fila_max["FechaHora"]):
        fecha_top = fila_max["FechaHora"]
    elif "Fecha" in fila_max and "Hora" in fila_max:
        try:
            fecha_top = datetime.combine(fila_max["Fecha"], fila_max["Hora"])
        except:
            fecha_top = fila_max["Fecha"]
    elif "Fecha" in fila_max:
        fecha_top = fila_max["Fecha"]

    st.markdown("## 🌎 Valor máximo registrado en Argentina")
    # --- Estilo visual con CSS ---
    st.markdown("""
    <style>
    div[data-testid="stMetricContainer"] {
        background: rgba(240, 248, 255, 0.6);
        border: 1px solid rgba(200, 200, 200, 0.3);
        border-radius: 12px;
        padding: 16px;
        text-align: center;
        box-shadow: 0 1px 6px rgba(0,0,0,0.1);
        transition: all 0.2s ease-in-out;
    }
    div[data-testid="stMetricContainer"]:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 10px rgba(0,0,0,0.15);
    }
    div[data-testid="stMetricLabel"] > div {
        font-size: 16px;
        font-weight: 600;
        color: #2E3B55;
    }
    div[data-testid="stMetricValue"] {
        font-size: 26px;
        font-weight: 700;
        color: #004aad;
    }
    </style>
    """, unsafe_allow_html=True)

    col1, col2, col3, col4 = st.columns([2.5, 1.5, 1.5, 1.5])
    col1.metric("Localidad", localidad_top)
    col2.metric("Resultado máximo V/m", f"{resultado_top:.2f}")
    col3.metric("Resultado máximo (%)", f"{resultado_top_pct:.2f}" if resultado_top_pct else "N/A")
    col4.metric("Fecha/Hora", str(fecha_top))

# ------------------- RESUMEN GENERAL DE LOCALIDADES ------------------
if "tabla_maestra" in st.session_state and not st.session_state["tabla_maestra"].empty:
    df = st.session_state["tabla_maestra"].copy()
    df["Resultado"] = pd.to_numeric(df["Resultado"], errors="coerce")
    
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], dayfirst=True, errors='coerce').dt.date
    if "Hora" in df.columns:
        df["Hora"] = pd.to_datetime(df["Hora"], errors='coerce').dt.time

    if "Fecha" in df.columns and "Hora" in df.columns:
        df["FechaHora"] = df.apply(lambda x: datetime.combine(x["Fecha"], x["Hora"]) if pd.notna(x["Fecha"]) and pd.notna(x["Hora"]) else pd.NaT, axis=1)
    else:
        df["FechaHora"] = pd.NaT

    resumen_localidad = []
    for (ccte, prov, loc), g in df.groupby(["CCTE","Provincia","Localidad"]):
        inicio = g["FechaHora"].min() if "FechaHora" in g.columns else None
        fin = g["FechaHora"].max() if "FechaHora" in g.columns else None
        tiempo_total_localidad = calcular_tiempo_total_por_archivo(g)
        max_res = g["Resultado"].max() if pd.notna(g["Resultado"].max()) else None
        resumen_localidad.append({
            "CCTE": ccte,
            "Provincia": prov,
            "Localidad": loc,
            "Inicio": inicio,
            "Fin": fin,
            "Mediciones": len(g),
            "Tiempo mediciones": format_timedelta_long(tiempo_total_localidad),
            "Resultado Max (V/m)": max_res,
            "Resultado Max (%)": max_res**2 / 3770 / 0.20021 * 100 if max_res else None,
            "N° Expediente": ", ".join(sorted(g["Expediente"].dropna().unique().astype(str))),
            "Sonda utilizada": ", ".join(sorted(g["Sonda"].dropna().unique().astype(str))) if "Sonda" in g.columns else "N/A"
        })

    resumen_localidad_df = pd.DataFrame(resumen_localidad)
    st.header("📊 Resumen general por localidad")
    st.dataframe(resumen_localidad_df)
if not resumen_localidad_df.empty:
    if st.button("📥 Exportar resumen por localidad a Excel"):
        try:
            # Crear archivo Excel
            ruta_excel = "resumen_localidades.xlsx"
            resumen_localidad_df.to_excel(ruta_excel, index=False)

            wb = openpyxl.load_workbook(ruta_excel)
            ws = wb.active

            # Insertar logo arriba
            try:
                logo_path = "logo_enacom.png"
                img = XLImage(logo_path)
                img.width, img.height = 200, 70
                ws.add_image(img, "A1")
                ws.insert_rows(1, amount=5)  # empujar tabla hacia abajo
            except Exception as e:
                st.warning(f"No se pudo insertar logo: {e}")

            # Aplicar estilos a títulos
            for cell in ws[6]:  # fila de encabezados después de empujar
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Centrar todas las celdas
            for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            wb.save(ruta_excel)
            st.success(f"Archivo '{ruta_excel}' generado con formato y logo.")
            with open(ruta_excel, "rb") as f:
                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=f,
                    file_name=ruta_excel,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        except Exception as e:
            st.error(f"Error exportando Excel: {e}")



#----------------------------- GRAFICOS-------------------------------------
if not st.session_state["tabla_maestra"].empty:
    df_grafico = st.session_state["tabla_maestra"].copy()
    
    # Distribución de puntos medidos por CCTE
    df_pie = df_grafico.groupby("CCTE").size().reset_index(name="Cantidad Puntos")
    fig_pie = px.pie(
        df_pie,
        names="CCTE",
        values="Cantidad Puntos",
        title="Distribución de puntos medidos por CCTE",
        color_discrete_sequence=px.colors.qualitative.Set3
    )

    # Localidades por Provincia y CCTE
    resumen = df_grafico.groupby(["Provincia","CCTE"])["Localidad"].nunique().reset_index(name="CantidadLocalidades")
    fig_bar = px.bar(
        resumen,
        x="Provincia",
        y="CantidadLocalidades",
        color="CCTE",
        text="CantidadLocalidades",
        barmode="group",
        title="Localidades por Provincia y CCTE"
    )

    st.subheader("📊 Resumen de mediciones y localidades")
    col1, col2 = st.columns(2)
    with col1:
        st.plotly_chart(fig_pie, use_container_width=True)
    with col2:
        st.plotly_chart(fig_bar, use_container_width=True)


# ---------------------- TABLA MAESTRA ----------------------------
st.header("📊 Tabla Maestra Actualizada")
if not st.session_state["tabla_maestra"].empty:
    df_maestra = st.session_state["tabla_maestra"].copy()
    df_maestra = df_maestra.dropna(axis=1, how="all")

    st.dataframe(df_maestra.reset_index(drop=True))

    if st.button("Exportar datos a Excel"):
        try:
            df_maestra.to_excel("tabla_maestra.xlsx", index=False)
            st.success("Archivo 'tabla_maestra.xlsx' generado correctamente.")
        except Exception as e:
            st.error(f"No se pudo exportar tabla maestra: {e}")
else:
    st.info("La tabla maestra está vacía. Cargá archivos a la izquierda.")

# ------------------- RESUMEN Y EDICIÓN DE LOCALIDAD ------------------
st.header("📊 Gestión de Localidades")

col1, col2, col3, col4 = st.columns([1, 1, 1, 1])

with col1:
    ccte_filtro = st.selectbox(
        "Filtrar CCTE",
        ["Todos"] + sorted(st.session_state["tabla_maestra"]["CCTE"].dropna().unique().tolist())
    )

    if ccte_filtro == "Todos":
        df_filtrado_ccte = st.session_state["tabla_maestra"].copy()
    else:
        df_filtrado_ccte = st.session_state["tabla_maestra"][
            st.session_state["tabla_maestra"]["CCTE"] == ccte_filtro
        ].copy()

with col2:
    # Selección de Provincia dentro del CCTE elegido
    provincia_filtro = st.selectbox(
        "Filtrar Provincia",
        ["Todas"] + sorted(df_filtrado_ccte["Provincia"].dropna().unique().tolist())
    )

    # Filtrado base según Provincia
    if provincia_filtro == "Todas":
        df_filtrado_prov = df_filtrado_ccte.copy()
    else:
        df_filtrado_prov = df_filtrado_ccte[df_filtrado_ccte["Provincia"] == provincia_filtro].copy()

# 🔧 Filtro por Año (incluye 2025 si existe en los datos)
with col4:
    año_filtro = "Todos"
    if not df_filtrado_prov.empty and "Fecha" in df_filtrado_prov.columns:
        # Convertimos Fecha a datetime para extraer año (acepta dd/mm/aaaa)
        _años = pd.to_datetime(df_filtrado_prov["Fecha"], dayfirst=True, errors="coerce").dt.year
        df_filtrado_prov["_Año"] = _años

        años_disponibles = (
            df_filtrado_prov["_Año"]
            .dropna()
            .astype(int)
            .unique()
            .tolist()
        )
        # Ordenamos descendente para que 2025 aparezca arriba si está
        años_disponibles = sorted(años_disponibles, reverse=True)

        # Armamos selectbox (si no hay años válidos, mostramos 'Todos' solamente)
        opciones_año = ["Todos"] + [str(a) for a in años_disponibles]
        año_filtro = st.selectbox("📅 Año", opciones_año, index=0)

        # Aplicamos filtro si corresponde
        if año_filtro != "Todos":
            df_filtrado_prov = df_filtrado_prov[df_filtrado_prov["_Año"] == int(año_filtro)]

        # Limpieza de la columna auxiliar para no “ensuciar” el DF
        if "_Año" in df_filtrado_prov.columns:
            df_filtrado_prov = df_filtrado_prov.drop(columns=["_Año"])

with col3:
    # Selección de Localidad dentro de la Provincia elegida
    localidades_cargadas = df_filtrado_prov["Localidad"].dropna().unique().tolist()
    localidad_seleccionada = st.selectbox("Seleccionar Localidad", [""] + sorted(localidades_cargadas))

    # Subset final
if localidad_seleccionada:
    df_localidad = df_filtrado_prov[df_filtrado_prov["Localidad"] == localidad_seleccionada].copy()
else:
    df_localidad = df_filtrado_prov.copy()

# Convertir fechas y horas
if "Fecha" in df_localidad.columns:
    df_localidad["Fecha"] = pd.to_datetime(df_localidad["Fecha"], dayfirst=True, errors='coerce').dt.date
if "Hora" in df_localidad.columns:
    df_localidad["Hora"] = pd.to_datetime(df_localidad["Hora"], errors='coerce').dt.time
if "Fecha" in df_localidad.columns and "Hora" in df_localidad.columns:
    df_localidad["FechaHora"] = df_localidad.apply(
        lambda x: datetime.combine(x["Fecha"], x["Hora"]) if pd.notna(x["Fecha"]) and pd.notna(x["Hora"]) else pd.NaT,
        axis=1
    )
else:
    df_localidad["FechaHora"] = pd.NaT    

# ---------------- Datos generales ----------------
if localidad_seleccionada:    
    provincia_real = df_localidad["Provincia"].iloc[0] if "Provincia" in df_localidad.columns else "N/A"
    titulo_scope = f"la localidad {localidad_seleccionada}, {provincia_real}"
elif provincia_filtro != "Todas":
    titulo_scope = f"{provincia_filtro}"
elif ccte_filtro != "Todos":
    titulo_scope = f"CCTE {ccte_filtro}"
else:
    titulo_scope = "todo el país"

st.subheader(f"Mediciones RNI de {titulo_scope}")


tiempo_total_localidad = calcular_tiempo_total_por_archivo(df_localidad)
total_puntos = len(df_localidad)
max_resultado = df_localidad["Resultado"].max() if "Resultado" in df_localidad.columns else None
max_resultado_pct = max_resultado**2 / 3770 / 0.20021 * 100 if pd.notna(max_resultado) else None
sondas = df_localidad["Sonda"].dropna().unique().tolist() if "Sonda" in df_localidad.columns else []

st.write(f"Cantidad total de puntos medidos: {total_puntos}")
st.write(f"Máximo Resultado (V/m): {max_resultado}")
st.write(f"Sonda utilizada: {', '.join(sondas) if sondas else 'N/A'}")
st.write(f"Tiempo total de mediciones: {format_timedelta_long(tiempo_total_localidad)} horas")

# ---------------- Resumen por día y mes ----------------
if "FechaHora" in df_localidad.columns and not df_localidad.empty:
    df_localidad["FechaHora"] = pd.to_datetime(df_localidad["FechaHora"])
    df_localidad["Fecha"] = df_localidad["FechaHora"].dt.date
    df_localidad["Mes"] = df_localidad["FechaHora"].dt.to_period("M")

    # --- Resumen diario ---
    def resumen_por_dia(df):
        tiempo_total = calcular_tiempo_total_por_archivo(df)
        hora_inicio = df["FechaHora"].min().strftime("%H:%M:%S")
        hora_fin = df["FechaHora"].max().strftime("%H:%M:%S")
        puntos = len(df)
        localidades = ", ".join(sorted(df["Localidad"].dropna().unique()))
        return pd.Series({
            "Hora de inicio": hora_inicio,
            "Hora de fin": hora_fin,
            "Tiempo total trabajado": format_timedelta_long(tiempo_total),
            "Cantidad de puntos medidos": puntos,
            "Localidades trabajadas (por día)": localidades
        })

    resumen_dias = df_localidad.groupby("Fecha").apply(resumen_por_dia).reset_index().rename(columns={"Fecha":"Fecha de medición"})
    
    # --- Resumen mensual ---
    resumen_mensual = df_localidad.groupby("Mes").agg({
        "FechaHora": ["min","max"],
        "Localidad": lambda x: ", ".join(sorted(x.dropna().unique())),
        "Resultado": "count"
    }).reset_index()

    resumen_mensual.columns = ["Mes","Hora inicio","Hora fin","Localidades trabajadas","Cantidad puntos"]

    # Calcular tiempo total trabajado por mes
    def calcular_tiempo_mes(g):
        return format_timedelta_long(calcular_tiempo_total_por_archivo(g))

    tiempo_por_mes = df_localidad.groupby("Mes").apply(calcular_tiempo_mes).reset_index(name="Horas trabajadas")
    resumen_mensual = resumen_mensual.merge(tiempo_por_mes, on="Mes")

 # -------- Tabs para elegir vista --------
    tab1, tab2, tab3 = st.tabs(["📅 Resumen Diario", "🗓️ Resumen Mensual", "📊 Gráfico"])

    with tab1:
        st.markdown(f"### ⏱️ Tiempo trabajado por día en {titulo_scope}")
        st.dataframe(resumen_dias)
    with tab2:
        st.markdown(f"### 📅 Mediciones Totales por mes en {titulo_scope}")
        st.dataframe(resumen_mensual)

    with tab3:
        if not resumen_mensual.empty:
            st.markdown(f"### 📊 Gráfico mensual de mediciones y tiempo trabajado en {titulo_scope}")

        # Crear figura con dos ejes: cantidad de puntos y horas trabajadas
        fig = go.Figure()

        # Barra: Cantidad de puntos
        fig.add_trace(go.Bar(
            x=resumen_mensual["Mes"].astype(str),
            y=resumen_mensual["Cantidad puntos"],
            name="Cantidad puntos",
            marker_color="steelblue",
            yaxis="y1",
            text=resumen_mensual["Cantidad puntos"],
            textposition="auto",
            hovertext=resumen_mensual["Localidades trabajadas"],  # 👈 tooltip
            hovertemplate="<b>%{x}</b><br>Puntos: %{y}<br>Localidades: %{hovertext}"
        ))

        # Línea: Horas trabajadas   
        def tiempo_a_horas(s):
            h, m, sec = map(int, s.split(":"))
            return h + m/60 + sec/3600

        resumen_mensual["Horas trabajadas num"] = resumen_mensual["Horas trabajadas"].apply(tiempo_a_horas)

        fig.add_trace(go.Scatter(
            x=resumen_mensual["Mes"].astype(str),
            y=resumen_mensual["Horas trabajadas num"],
            name="Horas trabajadas",
            yaxis="y2",
            mode="lines+markers",
            line=dict(color="orange", width=2)
        ))

        # Configuración de ejes
        fig.update_layout(
            xaxis=dict(title="Mes"),
            yaxis=dict(title="Cantidad de puntos", side="left"),
            yaxis2=dict(
                title="Horas trabajadas",
                overlaying="y",
                side="right"
            ),
            legend=dict(x=0.01, y=0.99),
            template="plotly_white",
            height=450
        )

        st.plotly_chart(fig, use_container_width=True)

# ---------------- Semáforo ----------------
if max_resultado_pct and not df_localidad.empty:
    rangos_colores = [
        (0, 1, "#84C2F5"), (1, 2, "#489DFF"), (2, 4, "#006BD6"),
        (4, 8, "#A9E7A9"), (8, 15, "#89DD89"), (15, 20, "#4D9623"),
        (20, 35, "#D9FF00"), (35, 50, "#F39A6D"), (50, 100, "#E68200"),
        (100, float("inf"), "#CC0000")
    ]
    color_localidad = next((color for low, high, color in rangos_colores if low <= max_resultado_pct < high), "#FFFFFF")
    st.markdown(
        f"""
        <div style="
            background-color:{color_localidad};
            padding:20px;
            border-radius:10px;
            text-align:center;
            font-size:24px;
            font-weight:bold;
            color:#000;">
            Resultado máximo en {titulo_scope}: {max_resultado_pct:.2f} %
        </div>
        """,
        unsafe_allow_html=True
    )
# Imagen del semáforo de colores 
    st.image("mapa de calor.png", caption="Escala de colores para interpretar los resultados", width='stretch')

# ------------------- MAPA INTERACTIVO ------------------
if "Lat" in df_localidad.columns and "Lon" in df_localidad.columns and not df_localidad.empty:
    coords = df_localidad.dropna(subset=["Lat", "Lon"])[["Lat", "Lon", "Localidad", "Resultado"]].copy()
    if not coords.empty:
        coords["lat"] = coords["Lat"].apply(lambda x: -abs(x))
        coords["lon"] = coords["Lon"].apply(lambda x: -abs(x))

        rangos_colores_map = [
            (0, 1, [132, 194, 245]),
            (1, 2, [72, 157, 255]),
            (2, 4, [0, 107, 214]),
            (4, 8, [169, 231, 169]),
            (8, 15, [137, 221, 137]),
            (15, 20, [77, 150, 35]),
            (20, 35, [217, 255, 0]),
            (35, 50, [243, 154, 109]),
            (50, 100, [230, 130, 0]),
            (100, float("inf"), [204, 0, 0])
        ]

        def color_semaforo(valor):
            if pd.isna(valor):
                return [200, 200, 200]
            for low, high, color in rangos_colores_map:
                if low <= valor < high:
                    return color
            return [0, 0, 0]

        coords["color"] = coords["Resultado"].apply(color_semaforo)

        st.subheader("🗺️ Mapa Semaforizado")
        mapa = pdk.Deck(
            map_style="https://basemaps.cartocdn.com/gl/positron-gl-style/style.json",
            initial_view_state=pdk.ViewState(
                latitude=coords["lat"].mean(),
                longitude=coords["lon"].mean(),
                zoom=6,
                pitch=0,
            ),
            layers=[
                pdk.Layer(
                    "ScatterplotLayer",
                    data=coords,
                    get_position='[lon, lat]',
                    get_fill_color='color',
                    get_radius=12,
                    pickable=True,
                )
            ],
            tooltip={"text": "Localidad: {Localidad}\nResultado: {Resultado}"}
        )
        st.pydeck_chart(mapa)

# -------------------- Edición de información (plegable) --------------------
if localidad_seleccionada:
    ultima_fecha = None
    if "FechaCarga" in st.session_state["tabla_maestra"].columns:
        mask_fecha = st.session_state["tabla_maestra"]["Localidad"] == localidad_seleccionada
        if mask_fecha.any():
            ultima_fecha = st.session_state["tabla_maestra"].loc[mask_fecha, "FechaCarga"].max()

    expander_title = f"✏️ Editar información de {localidad_seleccionada}"
    if ultima_fecha is not None and pd.notna(ultima_fecha):
        expander_title += f" (Última modificación: {ultima_fecha.strftime('%d/%m/%Y %H:%M:%S')})"

    with st.expander(expander_title, expanded=False):
        ccte_actual = df_localidad["CCTE"].iloc[0]
        provincia_actual = df_localidad["Provincia"].iloc[0]
        localidad_actual = df_localidad["Localidad"].iloc[0]
        expediente_actual = df_localidad["Expediente"].iloc[0]

        nuevo_ccte = st.selectbox(
            "CCTE",
            ["CABA", "Buenos Aires", "Comodoro Rivadavia", "Córdoba", "Neuquén", "Posadas", "Salta"],
            index=["CABA","Buenos Aires","Comodoro Rivadavia","Córdoba","Neuquén","Posadas","Salta"].index(ccte_actual)
        )
        nueva_provincia = st.selectbox(
            "Provincia",
            ["Buenos Aires","CABA","Catamarca","Chaco","Chubut","Córdoba","Corrientes","Entre Ríos","Formosa","Jujuy",
             "La Pampa","La Rioja","Mendoza","Misiones","Neuquén","Río Negro","Salta","San Juan","San Luis","Santa Cruz",
             "Santa Fe","Santiago del Estero","Tierra del Fuego","Tucumán"],
            index=["Buenos Aires","CABA","Catamarca","Chaco","Chubut","Córdoba","Corrientes","Entre Ríos","Formosa","Jujuy",
                   "La Pampa","La Rioja","Mendoza","Misiones","Neuquén","Río Negro","Salta","San Juan","San Luis","Santa Cruz",
                   "Santa Fe","Santiago del Estero","Tierra del Fuego","Tucumán"].index(provincia_actual)
        )
        nueva_localidad = st.text_input("Localidad", value=localidad_actual)
        nuevo_expediente = st.text_input("Expediente", value=expediente_actual)

        if "FechaCarga" not in st.session_state["tabla_maestra"].columns:
            st.session_state["tabla_maestra"]["FechaCarga"] = pd.NaT

        def guardar_cambios():
            mask = st.session_state["tabla_maestra"]["Localidad"] == localidad_actual
            st.session_state["tabla_maestra"].loc[mask, "CCTE"] = nuevo_ccte
            st.session_state["tabla_maestra"].loc[mask, "Provincia"] = nueva_provincia
            st.session_state["tabla_maestra"].loc[mask, "Localidad"] = nueva_localidad
            st.session_state["tabla_maestra"].loc[mask, "Expediente"] = nuevo_expediente
            st.session_state["tabla_maestra"].loc[mask, "FechaCarga"] = datetime.now()

            try:
                pickle.dump(st.session_state["tabla_maestra"], open(PICKLE_FILE, "wb"))
                st.success("Cambios guardados correctamente")
            except Exception as e:
                st.error(f"No se pudieron guardar los cambios: {e}")

        st.button("💾 Guardar cambios", on_click=guardar_cambios)

        def eliminar_localidad():
            mask = st.session_state["tabla_maestra"]["Localidad"] == localidad_actual
            if mask.any():
                st.session_state["tabla_maestra"] = st.session_state["tabla_maestra"].loc[~mask]
                try:
                    pickle.dump(st.session_state["tabla_maestra"], open(PICKLE_FILE, "wb"))
                    st.success(f"Localidad '{localidad_actual}' eliminada correctamente")
                    st.experimental_rerun()  # recarga la app para reflejar cambios
                except Exception as e:
                    st.error(f"No se pudo eliminar la localidad: {e}")
            else:
                st.warning("No se encontró la localidad para eliminar.")

        st.button("🗑️ Eliminar localidad", on_click=eliminar_localidad)
